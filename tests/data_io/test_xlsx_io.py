import pytest
from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook

from app.data_io.xlsx_io import read_sot_xlsx, read_tgt_xlsx
from app.data_io.xlsx_io import read_tgt_xlsx, write_tgt_xlsx


@pytest.fixture
def sot_path():
    return "tests/sample_input_files/SOT_sample.xlsx"


@pytest.fixture
def tgt_path():
    return "tests/sample_input_files/TGT_sample.xlsx"


def test_read_sot_xlsx_valid(sot_path):
    headers, rows = read_sot_xlsx(sot_path, sheet_name="SOT_Data")

    # sanity checks
    assert isinstance(headers, list)
    assert isinstance(rows, list)
    assert len(headers) > 0
    assert all(isinstance(r, dict) for r in rows)

    # required columns exist
    assert "REC ID" in headers
    assert any(r["REC ID"] == "REC-0123" for r in rows)


def test_read_tgt_xlsx_valid(tgt_path):
    wb, ws = read_tgt_xlsx(tgt_path, sheet_name="Sheet1")

    # workbook and sheet object
    assert wb is not None
    assert ws.title

    # header sanity
    first_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Record ID" in first_row
    assert "Record Name" in first_row


def test_read_sot_xlsx_duplicate_headers(tmp_path):
    # create temporary workbook with duplicate headers
    dup_file = tmp_path / "dup_headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "SOT_Data"
    ws.append(["A", "A"])
    ws.append(["1", "2"])
    wb.save(dup_file)

    with pytest.raises(ValueError, match="duplicate column names"):
        read_sot_xlsx(str(dup_file), sheet_name="SOT_Data")


def test_write_tgt_xlsx_preserves_fill(tmp_path, tgt_path):

    wb, ws = read_tgt_xlsx(tgt_path, sheet_name="Sheet1")
    orig_fill = ws.cell(row=2, column=1).fill

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    updated_rows = []
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2
    ):
        row_dict = dict(zip(headers, row))
        if r_idx == 2:
            row_dict["Record Name"] = "UPDATED VALUE"
        updated_rows.append(row_dict)

    out_file = write_tgt_xlsx(
        wb,
        ws,
        updated_rows,
        "TGT_sample.xlsx",
        tmp_path,
    )
    wb_new = load_workbook(out_file)
    ws_new = wb_new.active
    new_fill = ws_new.cell(row=2, column=1).fill

    # verify update applied
    assert (
        ws_new.cell(row=2, column=headers.index("Record Name") + 1).value
        == "UPDATED VALUE"
    )

    # only compare fills if original had a defined pattern
    if orig_fill.patternType:
        assert isinstance(new_fill, PatternFill)
        assert new_fill.patternType == orig_fill.patternType
        assert new_fill.start_color.rgb == orig_fill.start_color.rgb
