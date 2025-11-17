from __future__ import annotations
from typing import List, Dict, Tuple, Optional
from openpyxl import load_workbook

from config import OUTPUT_DIR


def read_sot_xlsx(
    file_path: str, sheet_name: str
) -> Tuple[List[str], List[Dict[str, str]]]:
    """
    Read SOT spreadsheet for data only (ignore styles).
    Returns (headers, list of dicts).
    """
    if not sheet_name:
        raise ValueError("SOT sheet name must be provided in config.py")

    wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    header_row = [
        str(c).strip() if c else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]

    headers = [h for h in header_row if h]
    if len(headers) != len(set(headers)):
        raise ValueError(f"{file_path}: duplicate column names detected in SOT.")

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = ["" if v is None else str(v).strip() for v in row[: len(headers)]]
        if any(values):
            data.append(dict(zip(headers, values)))
    return headers, data


def read_tgt_xlsx(file_path: str, sheet_name: str) -> Tuple:
    """
    Read TGT spreadsheet with format preservation (openpyxl workbook object).
    This allows future update of cell values while keeping fills, fonts, etc.
    """
    if not sheet_name:
        raise ValueError("TGT sheet name must be provided in config.py")

    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    return wb, ws


from datetime import datetime
from pathlib import Path


def write_tgt_xlsx(
    wb, ws, updated_rows, tgt_filename: str, output_dir: str = OUTPUT_DIR
) -> str:
    """
    Write updated TGT workbook while preserving styles (fills, fonts, etc.).
    Arguments:
      wb, ws: workbook and worksheet returned from read_tgt_xlsx()
      updated_rows: list of dicts with updated values (headers must match first row)
      output_dir: directory where output will be saved
      tgt_filename: base filename of original TGT file

    Returns path of the newly saved file.
    """
    headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {h: idx + 1 for idx, h in enumerate(headers)}

    # Update cells based on updated_rows (row 2 onward)
    for i, row_dict in enumerate(updated_rows, start=2):
        for col_name, new_value in row_dict.items():
            if col_name in header_index:
                cell = ws.cell(row=i, column=header_index[col_name])
                # Only update value; openpyxl keeps styles automatically
                cell.value = new_value

    # generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_name = f"{Path(tgt_filename).stem}_updated_{timestamp}.xlsx"
    out_path = Path(output_dir) / out_name
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return str(out_path)


# ! test only: To inspect the output of the xlsx TGT file to make sure styles are preserved
if __name__ == "__main__":
    tgt_file = "tests/sample_input_files/TGT_sample.xlsx"

    wb, ws = read_tgt_xlsx(tgt_file)

    # read headers correctly
    headers = [
        str(c).strip()
        for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]

    # build current rows
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append(dict(zip(headers, row)))

    # add one new record (new REC ID)
    new_record = {h: "" for h in headers}
    new_record.update(
        {
            "REC ID": "REC-9999",
            "REC Name": "Manual Test New Entry",
            "Description": "Added for manual verification of fill retention",
            "Owner": "Tester (00000000)",
            "Status": "Active",
        }
    )
    rows.append(new_record)

    out_path = write_tgt_xlsx(wb, ws, rows, "TGT_sample.xlsx", ".")
    print(f"New file written to: {out_path}")
